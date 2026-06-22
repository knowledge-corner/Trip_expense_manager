"""Excel helpers for full-database backup/restore and per-trip expense bulk import."""

from datetime import datetime, date
from decimal import Decimal

from django.contrib.auth import get_user_model
from openpyxl import Workbook, load_workbook

from core.categories import EXPENSE_CATEGORY_MAP
from expenses.models import Expense, ExpenseSplit
from reviews.models import PlaceReview
from trips.models import CategoryBudget, Hotel, ItineraryDay, Trip, TripParticipant

User = get_user_model()


# ---------------------------------------------------------------------------
# Full database backup (used before a migration, e.g. SQLite -> MySQL)
# ---------------------------------------------------------------------------

def build_backup_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    trips_ws = wb.create_sheet("Trips")
    trips_ws.append(
        ["id", "name", "description", "destination", "state", "start_date", "end_date",
         "budget", "default_split_type", "created_by_username", "created_at"]
    )
    for t in Trip.objects.all():
        trips_ws.append(
            [t.id, t.name, t.description, t.destination, t.state,
             t.start_date.isoformat() if t.start_date else "",
             t.end_date.isoformat() if t.end_date else "",
             float(t.budget) if t.budget is not None else "",
             t.default_split_type, t.created_by.username if t.created_by_id else "",
             t.created_at.isoformat() if t.created_at else ""]
        )

    participants_ws = wb.create_sheet("Participants")
    participants_ws.append(["id", "trip_id", "name", "user_username"])
    for p in TripParticipant.objects.all():
        participants_ws.append([p.id, p.trip_id, p.name, p.user.username if p.user_id else ""])

    hotels_ws = wb.create_sheet("Hotels")
    hotels_ws.append(
        ["id", "trip_id", "name", "location", "check_in_date", "check_out_date",
         "meal_plan", "booking_status", "notes"]
    )
    for h in Hotel.objects.all():
        hotels_ws.append(
            [h.id, h.trip_id, h.name, h.location,
             h.check_in_date.isoformat() if h.check_in_date else "",
             h.check_out_date.isoformat() if h.check_out_date else "",
             h.meal_plan, h.booking_status, h.notes]
        )

    itinerary_ws = wb.create_sheet("ItineraryDays")
    itinerary_ws.append(["id", "trip_id", "date", "event", "hotel_id", "notes", "order"])
    for day in ItineraryDay.objects.all():
        itinerary_ws.append(
            [day.id, day.trip_id, day.date.isoformat(), day.event, day.hotel_id or "",
             day.notes, day.order]
        )

    budgets_ws = wb.create_sheet("CategoryBudgets")
    budgets_ws.append(["id", "trip_id", "category", "allocated_amount"])
    for b in CategoryBudget.objects.all():
        budgets_ws.append([b.id, b.trip_id, b.category, float(b.allocated_amount)])

    expenses_ws = wb.create_sheet("Expenses")
    expenses_ws.append(
        ["id", "trip_id", "category", "amount", "date", "description", "location",
         "hotel_id", "paid_by_id", "split_type", "created_by_username", "created_at"]
    )
    for e in Expense.objects.all():
        expenses_ws.append(
            [e.id, e.trip_id, e.category, float(e.amount), e.date.isoformat(),
             e.description, e.location, e.hotel_id or "", e.paid_by_id or "", e.split_type,
             e.created_by.username if e.created_by_id else "",
             e.created_at.isoformat() if e.created_at else ""]
        )

    splits_ws = wb.create_sheet("ExpenseSplits")
    splits_ws.append(["id", "expense_id", "participant_id", "amount"])
    for s in ExpenseSplit.objects.all():
        splits_ws.append([s.id, s.expense_id, s.participant_id, float(s.amount)])

    reviews_ws = wb.create_sheet("Reviews")
    reviews_ws.append(
        ["id", "trip_id", "place_name", "place_type", "location", "hotel_id", "rating",
         "amount_spent", "review_text", "would_revisit", "alternative_suggestion",
         "created_by_username", "created_at"]
    )
    for r in PlaceReview.objects.all():
        reviews_ws.append(
            [r.id, r.trip_id, r.place_name, r.place_type, r.location, r.hotel_id or "", r.rating,
             float(r.amount_spent) if r.amount_spent is not None else "",
             r.review_text, r.would_revisit, r.alternative_suggestion,
             r.created_by.username if r.created_by_id else "",
             r.created_at.isoformat() if r.created_at else ""]
        )

    return wb


def restore_backup_workbook(wb: Workbook, clear_existing=False):
    """Restore data from a workbook produced by build_backup_workbook().

    Existing primary keys are preserved so relations stay intact. Safe to run
    against an empty freshly-migrated database (e.g. after switching to MySQL).
    """
    if clear_existing:
        ExpenseSplit.objects.all().delete()
        Expense.objects.all().delete()
        PlaceReview.objects.all().delete()
        ItineraryDay.objects.all().delete()
        CategoryBudget.objects.all().delete()
        Hotel.objects.all().delete()
        TripParticipant.objects.all().delete()
        Trip.objects.all().delete()

    summary = {
        "trips": 0, "participants": 0, "hotels": 0, "itinerary_days": 0,
        "category_budgets": 0, "expenses": 0, "splits": 0, "reviews": 0,
    }

    def _date(v):
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
        return datetime.strptime(str(v), "%Y-%m-%d").date() if v else None

    def _get_user(username):
        if not username:
            return None
        user, _ = User.objects.get_or_create(username=username, defaults={"is_active": True})
        return user

    ws = wb["Trips"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        (tid, name, desc, dest, state, start, end, budget, split_type, owner, _created) = row
        Trip.objects.update_or_create(
            id=tid,
            defaults=dict(
                name=name, description=desc or "", destination=dest or "", state=state or "",
                start_date=_date(start), end_date=_date(end) if end else None,
                budget=Decimal(str(budget)) if budget not in (None, "") else None,
                default_split_type=split_type or Trip.SPLIT_EQUAL,
                created_by=_get_user(owner),
            ),
        )
        summary["trips"] += 1

    ws = wb["Participants"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        pid, trip_id, name, username = row
        TripParticipant.objects.update_or_create(
            id=pid, defaults=dict(trip_id=trip_id, name=name, user=_get_user(username))
        )
        summary["participants"] += 1

    if "Hotels" in wb.sheetnames:
        ws = wb["Hotels"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            (hid, trip_id, name, location, check_in, check_out, meal_plan,
             booking_status, notes) = row
            Hotel.objects.update_or_create(
                id=hid,
                defaults=dict(
                    trip_id=trip_id, name=name, location=location or "",
                    check_in_date=_date(check_in) if check_in else None,
                    check_out_date=_date(check_out) if check_out else None,
                    meal_plan=meal_plan or Hotel.MEAL_NONE,
                    booking_status=booking_status or Hotel.STATUS_PLANNED,
                    notes=notes or "",
                ),
            )
            summary["hotels"] += 1

    if "ItineraryDays" in wb.sheetnames:
        ws = wb["ItineraryDays"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            iid, trip_id, idate, event, hotel_id, notes, order = row
            ItineraryDay.objects.update_or_create(
                id=iid,
                defaults=dict(
                    trip_id=trip_id, date=_date(idate), event=event or "",
                    hotel_id=hotel_id or None, notes=notes or "", order=order or 0,
                ),
            )
            summary["itinerary_days"] += 1

    if "CategoryBudgets" in wb.sheetnames:
        ws = wb["CategoryBudgets"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            bid, trip_id, category, allocated_amount = row
            CategoryBudget.objects.update_or_create(
                id=bid,
                defaults=dict(
                    trip_id=trip_id, category=category,
                    allocated_amount=Decimal(str(allocated_amount)),
                ),
            )
            summary["category_budgets"] += 1

    ws = wb["Expenses"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    has_hotel_col = "hotel_id" in header
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        if has_hotel_col:
            (eid, trip_id, category, amount, edate, desc, location, hotel_id, paid_by_id,
             split_type, username, _created) = row
        else:
            (eid, trip_id, category, amount, edate, desc, location, paid_by_id,
             split_type, username, _created) = row
            hotel_id = None
        Expense.objects.update_or_create(
            id=eid,
            defaults=dict(
                trip_id=trip_id, category=category, amount=Decimal(str(amount)),
                date=_date(edate), description=desc or "", location=location or "",
                hotel_id=hotel_id or None,
                paid_by_id=paid_by_id or None, split_type=split_type or Expense.SPLIT_SINGLE,
                created_by=_get_user(username),
            ),
        )
        summary["expenses"] += 1

    ws = wb["ExpenseSplits"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        sid, expense_id, participant_id, amount = row
        ExpenseSplit.objects.update_or_create(
            id=sid, defaults=dict(expense_id=expense_id, participant_id=participant_id,
                                   amount=Decimal(str(amount)))
        )
        summary["splits"] += 1

    if "Reviews" in wb.sheetnames:
        ws = wb["Reviews"]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        has_hotel_col = "hotel_id" in header
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            if has_hotel_col:
                (rid, trip_id, place_name, place_type, location, hotel_id, rating,
                 amount_spent, review_text, would_revisit, alt, username, _created) = row
            else:
                (rid, trip_id, place_name, place_type, location, rating, amount_spent,
                 review_text, would_revisit, alt, username, _created) = row
                hotel_id = None
            PlaceReview.objects.update_or_create(
                id=rid,
                defaults=dict(
                    trip_id=trip_id, place_name=place_name, place_type=place_type or "other",
                    location=location or "", hotel_id=hotel_id or None, rating=rating or 5,
                    amount_spent=Decimal(str(amount_spent)) if amount_spent not in (None, "") else None,
                    review_text=review_text or "", would_revisit=bool(would_revisit),
                    alternative_suggestion=alt or "", created_by=_get_user(username),
                ),
            )
            summary["reviews"] += 1

    return summary


# ---------------------------------------------------------------------------
# Per-trip expense bulk import (everyday use: load historical expenses)
# ---------------------------------------------------------------------------

EXPENSE_IMPORT_HEADERS = ["date", "category", "amount", "description", "location", "paid_by", "split_type"]


def build_expense_import_template() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    ws.append(EXPENSE_IMPORT_HEADERS)
    ws.append(["2026-01-01", "food", 500, "Lunch at restaurant", "Mumbai", "Alice", "equal"])
    notes = wb.create_sheet("Valid Categories")
    notes.append(["value", "label"])
    for value, label in EXPENSE_CATEGORY_MAP.items():
        notes.append([value, label])
    return wb


ITINERARY_IMPORT_HEADERS = ["date", "day", "travel", "stay", "activities"]
ITINERARY_SHEET_ALIASES = ("Itinerary", "Itenary")


def build_itinerary_import_template() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Itinerary"
    ws.append(ITINERARY_IMPORT_HEADERS)
    ws.append(["2026-01-25", "Sunday", "Mumbai - Port Blair", "Bay Leaf Inn, Port Blair", "Cellular Jail\nCorbyn's Cove Beach"])
    return wb


def import_trip_itinerary(trip: Trip, wb: Workbook, created_by=None):
    """Import a day-wise itinerary (Date, Day, Travel, Stay, Activities) like the
    Google Sheets the user already plans trips with. 'Travel' becomes the day's
    event, 'Stay' + 'Activities' are combined into notes since hotel names in
    these sheets are free-text and not reliably mappable to a Hotel record.

    Sheet name and header row position vary across real sheets (e.g. "Itenarary",
    "Sheet1" with a blank first row), so both are detected rather than assumed.
    """
    sheet_name = None
    for name in wb.sheetnames:
        if "itiner" in name.lower() or "itenar" in name.lower():
            sheet_name = name
            break
    ws = wb[sheet_name] if sheet_name else wb.active

    header_row_idx, header = None, None
    for row in ws.iter_rows(min_row=1, max_row=10):
        values = [str(c.value).strip().lower() if c.value else "" for c in row]
        if "date" in values:
            header_row_idx = row[0].row
            header = values
            break
    if header is None:
        return 0, ["could not find a header row with a 'Date' column"]

    created, errors = 0, []
    existing_max_order = ItineraryDay.objects.filter(trip=trip).count()
    for idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        if not row or all(v is None for v in row):
            continue
        data = dict(zip(header, row))
        try:
            idate = data.get("date")
            if isinstance(idate, datetime):
                idate = idate.date()
            elif isinstance(idate, str):
                idate = datetime.strptime(idate, "%Y-%m-%d").date()
            if idate is None:
                raise ValueError("missing date")
            travel = str(data.get("travel") or data.get("event") or "").strip()
            stay = str(data.get("stay") or data.get("hotel") or "").strip()
            activities = str(data.get("activities") or data.get("itinerary") or data.get("notes") or "").strip()
            notes_parts = []
            if stay:
                notes_parts.append(f"Stay: {stay}")
            if activities:
                notes_parts.append(activities)
            ItineraryDay.objects.create(
                trip=trip, date=idate, event=travel,
                notes="\n".join(notes_parts), order=existing_max_order + created,
            )
            created += 1
        except Exception as exc:
            errors.append(f"row {idx}: {exc}")
    return created, errors


def import_trip_expenses(trip: Trip, wb: Workbook, created_by):
    ws = wb.active if wb.active.title == "Expenses" else wb["Expenses"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header = [str(h).strip().lower() if h else "" for h in header]
    created, errors = 0, []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None for v in row):
            continue
        data = dict(zip(header, row))
        try:
            category = str(data.get("category", "")).strip().lower()
            if category not in EXPENSE_CATEGORY_MAP:
                raise ValueError(f"unknown category '{category}'")
            edate = data.get("date")
            if isinstance(edate, datetime):
                edate = edate.date()
            elif isinstance(edate, str):
                edate = datetime.strptime(edate, "%Y-%m-%d").date()
            amount = Decimal(str(data.get("amount")))
            paid_by = None
            paid_by_name = (data.get("paid_by") or "").strip()
            if paid_by_name:
                paid_by, _ = TripParticipant.objects.get_or_create(trip=trip, name=paid_by_name)
            split_type = (data.get("split_type") or Expense.SPLIT_SINGLE).strip().lower()
            if split_type not in dict(Expense.SPLIT_TYPE_CHOICES):
                split_type = Expense.SPLIT_SINGLE
            expense = Expense.objects.create(
                trip=trip, category=category, amount=amount, date=edate,
                description=(data.get("description") or "").strip(),
                location=(data.get("location") or "").strip(),
                paid_by=paid_by, split_type=split_type, created_by=created_by,
            )
            if split_type == Expense.SPLIT_EQUAL and trip.participants.exists():
                participants = list(trip.participants.all())
                share = (amount / len(participants)).quantize(Decimal("0.01"))
                remainder = amount - share * len(participants)
                for i, participant in enumerate(participants):
                    ExpenseSplit.objects.create(
                        expense=expense, participant=participant,
                        amount=share + remainder if i == 0 else share,
                    )
            created += 1
        except Exception as exc:
            errors.append(f"row {idx}: {exc}")
    return created, errors
