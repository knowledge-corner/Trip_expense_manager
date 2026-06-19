from datetime import datetime

from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import load_workbook

from trips.models import Trip

from .excel import (
    build_backup_workbook,
    build_expense_import_template,
    import_trip_expenses,
    restore_backup_workbook,
)


@staff_member_required
def backup_page(request):
    return render(request, "core/backup.html")


@staff_member_required
def backup_download(request):
    wb = build_backup_workbook()
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"trip_expense_backup_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@staff_member_required
def backup_restore(request):
    if request.method == "POST" and request.FILES.get("backup_file"):
        wb = load_workbook(request.FILES["backup_file"])
        summary = restore_backup_workbook(wb, clear_existing=bool(request.POST.get("clear_existing")))
        messages.success(
            request,
            f"Restored: {summary['trips']} trips, {summary['participants']} participants, "
            f"{summary['expenses']} expenses, {summary['splits']} splits, {summary['reviews']} reviews.",
        )
    else:
        messages.error(request, "Please choose a backup .xlsx file to restore.")
    return redirect("core:backup_page")


@login_required
def expense_import_template(request):
    wb = build_expense_import_template()
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="expense_import_template.xlsx"'
    wb.save(response)
    return response


@login_required
def expense_import(request, trip_id):
    trip = get_object_or_404(Trip, pk=trip_id)
    if request.method == "POST" and request.FILES.get("expense_file"):
        wb = load_workbook(request.FILES["expense_file"])
        created, errors = import_trip_expenses(trip, wb, request.user)
        if created:
            messages.success(request, f"Imported {created} expenses.")
        if errors:
            messages.warning(request, "Some rows could not be imported: " + "; ".join(errors[:10]))
    else:
        messages.error(request, "Please choose an .xlsx file to import.")
    return redirect("trips:trip_detail", pk=trip_id)
